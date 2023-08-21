#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# SPDX-License-Identifier: GPL-2.0-only
#
# Compare table generator for IP flatten flow
#
# Copyright (C) 2023 Yeh, Hsin-Hsien <yhh76227@gmail.com>
#

import argparse 
import gzip
import importlib
import os
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule

from .utils.common import PKG_VERSION

VERSION = f"ip_flat_cmp version 1.0.0 ({PKG_VERSION})"

### Global Setting ###   {{{

IP_FLAT_RPT_NAME = 'ip_boundary_timing.brief'

CORNER_MAP = {
    'NORMAL_FFGNP_0p88v_125c_CBEST_125c_HOLD_'      : '0p88v_cbest_125c_hold',
    'NORMAL_FFGNP_0p88v_125c_CWORST_125c_HOLD_'     : '0p88v_cworst_125c_hold',
    'NORMAL_FFGNP_0p88v_125c_RCBEST_125c_HOLD_'     : '0p88v_rcbest_125c_hold',
    'NORMAL_FFGNP_0p88v_125c_RCWORST_125c_HOLD_'    : '0p88v_rcworst_125c_hold',
    'NORMAL_FFGNP_0p88v_m40c_CBEST_m40c_HOLD_'      : '0p88v_cbest_m40c_hold',
    'NORMAL_FFGNP_0p88v_m40c_CWORST_m40c_HOLD_'     : '0p88v_cworst_m40c_hold',
    'NORMAL_FFGNP_0p88v_m40c_RCBEST_m40c_HOLD_'     : '0p88v_rcbest_m40c_hold',
    'NORMAL_FFGNP_0p88v_m40c_RCWORST_m40c_HOLD_'    : '0p88v_rcworst_m40c_hold',
    'NORMAL_SSGNP_0p72v_125c_CWORST_125c_HOLD_'     : '0p72v_cworst_125c_hold',
    'NORMAL_SSGNP_0p72v_125c_CWORST_T_125c_SETUP_'  : '0p72v_cworst_t_125c_setup',
    'NORMAL_SSGNP_0p72v_125c_RCWORST_125c_HOLD_'    : '0p72v_rcworst_125c_hold',
    'NORMAL_SSGNP_0p72v_125c_RCWORST_T_125c_SETUP_' : '0p72v_rcworst_t_125c_setup',
    'NORMAL_SSGNP_0p72v_m40c_CWORST_m40c_HOLD_'     : '0p72v_cworst_m40c_hold',
    'NORMAL_SSGNP_0p72v_m40c_CWORST_T_m40c_SETUP_'  : '0p72v_cworst_t_m40c_setup',
    'NORMAL_SSGNP_0p72v_m40c_RCWORST_m40c_HOLD_'    : '0p72v_rcworst_m40c_hold',
    'NORMAL_SSGNP_0p72v_m40c_RCWORST_T_m40c_SETUP_' : '0p72v_rcworst_t_m40c_setup',
}

STATUS_FORMULA = {
    'total_chk' : '=IF(OR(C{0}<>"", D{0}<>"", $E{0}<>"", F{0}<>""), 1, 0)',
    'fatal_chk' : '=IF(AND(H{0}="inf", I{0}="inf"), "na", ' +
                   'IF(AND(H{0}="inf", I{0}<>"inf"), "", ' +
                   'IF(H{0}<0, "x", ' +
                   'IF(OR(I{0}="inf", AND(I{0}>=0, J{0}>-0.001)), "v", ""))))',
    'high_warn' : '=IF(C{0}<>"", "", ' +
                   'IF(AND(J{0}<>"na", OR(J{0}<=-0.1, I{0}<0)), "x", ""))',
    'low_warn'  : '=IF(OR(C{0}<>"", D{0}<>""), "", ' +
                   'IF(AND(J{0}<>"na", J{0}>-0.1, J{0}<=-0.001), "x", ' +
                   'IF(AND(J{0}<>"na", J{0}>=0.1), "c", "")))',
    'user_chk'  : ''
}

FATAL_CHK = [
    lambda post_slk: post_slk != 'inf' and post_slk < 0
]

HELP_MSG = (
    "Fatal Error : (post-flatten slack < 0ps)",
    "High Warn : (slack difference <= -100ps) or (pre-flatten slack < 0ps)",
    "Low Warn(x) : (-100ps < slack difference <= -0.001ps)",
    "Low Warn(c) : (slack difference >= 100ps), maybe need check",
)

#}}}

### Parameter ###   {{{

CS_THP, CS_CKG = range(2)
PT_ARL, PT_SLK, PT_CKG = range(3)
UC_THP, UC_STS, UC_COM = range(1, 4)
UCT_COR, UCT_PID, UCT_THP = range(3)
UT_STS, UT_COM = range(2)

BOLD_FONT1 = Font(bold=True)
RED_FONT1 = Font(color='ffcc0000')
GREEN_FONT1 = Font(color='ff006600')
YELLOW_FONT1 = Font(color='ff996600')

RED_FILL1 = PatternFill(fill_type='solid', start_color='ffffcccc')
PURPLE_FILL1 = PatternFill(fill_type='solid', start_color='ffdedce6')
GREEN_FILL1 = PatternFill(fill_type='solid', start_color='ff92d050')
GREEN_FILL2 = PatternFill(fill_type='solid', start_color='ffccffcc')
YELLOW_FILL1 = PatternFill(fill_type='solid', start_color='ffffffcc')

THIN_BLACK_SIDE = Side(border_style='thin', color='ff000000')
THIN_GREY_SIDE = Side(border_style='thin', color='ff808080')
AR_BORDER1 = Border(left=THIN_BLACK_SIDE, right=THIN_BLACK_SIDE, top=THIN_BLACK_SIDE, bottom=THIN_BLACK_SIDE)
AR_BORDER2 = Border(left=THIN_GREY_SIDE, right=THIN_GREY_SIDE, top=THIN_GREY_SIDE, bottom=THIN_GREY_SIDE)

LC_ALIGN = Alignment(horizontal='left', vertical='center', wrapText=True)
CC_ALIGN = Alignment(horizontal='center', vertical='center', wrapText=True)
RC_ALIGN = Alignment(horizontal='right', vertical='center', wrapText=True)

STS_CHK = CellIsRule(operator='==', formula=[1], font=GREEN_FONT1, fill=GREEN_FILL2)
STS_WAIT = CellIsRule(operator='==', formula=[0], font=RED_FONT1, fill=RED_FILL1)
CHK_FAIL = CellIsRule(operator='==', formula=['"x"'], font=RED_FONT1, fill=RED_FILL1)
CHK_RSV = CellIsRule(operator='==', formula=['"c"'], font=YELLOW_FONT1, fill=YELLOW_FILL1)

#}}}

### Function ###

def load_cfg(cfg_fn: str):
    """Load Tool Config"""  #{{{
    global CORNER_MAP
    global STATUS_FORMULA
    global FATAL_CHK
    global HELP_MSG

    cfg_mod = Path(cfg_fn).stem

    sys.path.insert(0, '')
    try:
        config = importlib.import_module(cfg_mod)
    except ModuleNotFoundError:
        print(f"ModuleNotFoundError: Please create '{cfg_mod}' module in current directory")
        exit(1)

    try:
        CORNER_MAP = config.CORNER_MAP
    except AttributeError:
        pass

    try:
        STATUS_FORMULA = config.STATUS_FORMULA
    except AttributeError:
        pass

    try:
        FATAL_CHK = config.FATAL_CHK
    except AttributeError:
        pass

    try:
        HELP_MSG = config.HELP_MSG
    except AttributeError:
        pass
#}}}

def load_timing_report(rpt_fp, col_sz: list) -> dict:
    """Load Timing Report"""  #{{{
    ## return: report table
    table = {}
    with open(rpt_fp, 'r') as f:
        for line in f:
            line = line.strip()

            if line.startswith('#'):
                pass
            elif line != '':
                tok_list = line.split(',')
                thpoint = tok_list[0].strip()
                arrival = tok_list[1].strip()
                slack = tok_list[3].strip()
                ckgroup = tok_list[5].strip()

                if arrival == 'INFINITY' or arrival == 'NA':
                    arrival_f = "inf"
                else:
                    arrival_f = float(arrival)

                if slack == 'INFINITY' or slack == 'NA':
                    slack_f = "inf"
                else:
                    slack_f = float(slack)

                table[thpoint] = (arrival_f, slack_f, ckgroup)  # table structure

                size = len(thpoint)
                if size > col_sz[CS_THP]:
                    col_sz[CS_THP] = size

                size = len(ckgroup)
                if size > col_sz[CS_CKG]:
                    col_sz[CS_CKG] = size

    return table
#}}}

def report_dump_xls(ws: openpyxl.worksheet, post_rpt: dict, pre_rpt: dict, corner: str) -> int:
    """Dump Compare Table"""  #{{{
    ## return: fatal count
    post_cnt, pre_cnt = len(post_rpt), len(pre_rpt)
    if post_cnt != pre_cnt:
        print("WARNING: Number of report paths is not equal.")
        print("INFO: Corner: {}".format(corner))
        print("INFO: (Post, Pre) = ({}, {})".format(post_cnt, pre_cnt))

    rid = 1
    status = ('ID', 'Total\nCheck', 'Fatal\nCheck', 'High\nWarn', 'Low\nWarn', 'User\nCheck')

    status_cnt = len(status)
    for cid in range(1, status_cnt+1):
        cell = ws.cell(rid, cid)
        cell.fill = PURPLE_FILL1
        cell.alignment = CC_ALIGN
        cell.border = AR_BORDER1 

    data_st = rid + 2 
    data_ed = post_cnt + 2 

    ## status summation

    ws[f'A{rid}'] = f'=A{data_ed}+1'
    ws[f'B{rid}'] = f'=SUM(B{data_st}:B{data_ed})'
    ws[f'C{rid}'] = f'=COUNTIF(C{data_st}:C{data_ed}, "="&"x")'
    ws[f'D{rid}'] = f'=COUNTIF(D{data_st}:D{data_ed}, "="&"x")'
    ws[f'E{rid}'] = f'=COUNTIF(E{data_st}:E{data_ed}, "="&"x")'
    ws[f'F{rid}'] = f'=COUNTIF(F{data_st}:F{data_ed}, "="&"x")'

    ## add table title

    rid += 1
    for cid, title in enumerate(status, 1):
        cell = ws.cell(rid, cid, title)
        cell.fill = GREEN_FILL1
        cell.border = AR_BORDER1 
        cell.alignment = CC_ALIGN

    ws.row_dimensions[rid].height = 24

    ws.column_dimensions['H'].width = 10 
    ws.column_dimensions['I'].width = 10 
    ws.column_dimensions['J'].width = 10 
    ws.column_dimensions['K'].width = 10 
    ws.column_dimensions['G'].width = 24 
    ws.column_dimensions['L'].width = 30 
    ws.column_dimensions['M'].width = 30 
    ws.column_dimensions['N'].width = 30 

    items = ('Throughpoint', 'Slack\npostFlat', 'Slack\npreFlat', 'Slack\nDiff', 
             'Group\nSame', 'Dir/CKin/CKout (postFlat)', 'Dir/CKin/CKout (preFlat)', 'Comment')

    for cid, title in enumerate(items, start=7):
        cell = ws.cell(rid, cid, title)
        cell.fill = GREEN_FILL1
        cell.border = AR_BORDER1
        cell.alignment = LC_ALIGN

    ## add status condition format

    ws.conditional_formatting.add(f'B{data_st}:B{data_ed}',
            CellIsRule(operator='==', formula=[1], font=GREEN_FONT1, fill=GREEN_FILL2))
    ws.conditional_formatting.add(f'B{data_st}:B{data_ed}',
            CellIsRule(operator='==', formula=[0], font=RED_FONT1, fill=RED_FILL1))
    ws.conditional_formatting.add(f'C{data_st}:F{data_ed}',
            CellIsRule(operator='==', formula=['"x"'], font=RED_FONT1, fill=RED_FILL1))
    ws.conditional_formatting.add(f'C{data_st}:F{data_ed}',
            CellIsRule(operator='==', formula=['"c"'], font=YELLOW_FONT1, fill=YELLOW_FILL1))
    ws.conditional_formatting.add(f'H{data_st}:J{data_ed}',
            CellIsRule(operator='<', formula=[0.0], font=RED_FONT1, fill=RED_FILL1))
    ws.conditional_formatting.add(f'K{data_st}:K{data_ed}',
            CellIsRule(operator='==', formula=['"x"'], font=RED_FONT1, fill=RED_FILL1))

    ## add contents

    fatal_cnt = 0
    rid += 1
    for thp, post_val in post_rpt.items():
        try:
            pre_val = pre_rpt[thp]
        except KeyError:
            print("WARNING: Cannot find the path in pre-flatten report")
            print("INFO: Corner: {}".format(corner))
            print("INFO: Throughpoint: {}".format(thp))
            pre_val = (thp, "inf", "inf", '<NA|NA|NA>')

        value_list = (rid - data_st, 
                      STATUS_FORMULA['total_chk'].format(rid), 
                      STATUS_FORMULA['fatal_chk'].format(rid), 
                      STATUS_FORMULA['high_warn'].format(rid), 
                      STATUS_FORMULA['low_warn'].format(rid), 
                      STATUS_FORMULA['user_chk'].format(rid), 
                      thp,
                      post_val[PT_SLK],
                      pre_val[PT_SLK],
                      '=IF(OR(H{0}="inf", I{0}="inf"), "na", H{0}-I{0})'.format(rid),
                      'o' if post_val[PT_CKG] == pre_val[PT_CKG] else 'x',
                      post_val[PT_CKG],
                      pre_val[PT_CKG], ' ')

        for fatal_func in FATAL_CHK:
            if fatal_func(post_val[PT_SLK]):
                fatal_cnt += 1
                break

        for cid, value in enumerate(value_list, start=1): 
            cell = ws.cell(rid, cid, value)
            cell.border = AR_BORDER2
            if cid <= 6 or cid == 11:
                cell.alignment = CC_ALIGN
            elif 8 <= cid <= 10:
                cell.alignment = RC_ALIGN
                cell.number_format = "0.0000"

        rid += 1

    ## add filter

    ws.auto_filter.ref = f"A{data_st-1}:M{data_ed}"
    print(f"[INFO] Corner: {corner} (import done)")
    return fatal_cnt
#}}}

def load_ucheck_cfg(user_fp) -> dict:
    """Load User Check Config"""  #{{{
    # return: user check table
    #  
    # User Table Structure:
    #
    # table -+- pathA -+- cornerA -+- <id0>
    #        |         |           +- <id1>
    #        |         |           +- ...
    #        |         |           +- [all]
    #        |         +- all -----+- <id2>
    #        |                     +- ...
    #        |                     +- [all]
    #        +- pathB -+- cornerB -+- <id0>
    #                  |           +- <id1>
    #                  |           +- ...
    #                  |           +- [all]
    #                  +- all -----+- <id2>
    #                              +- ...
    #                              +- [all]
    #
    uchk_re = re.compile(r"\"([^\"]+)\", \"([^\"]+)\", \"([^\"]+)\"")
    table = {}
    
    with open(user_fp, 'r') as f:
        for fno, line in enumerate(f.readlines(), start=1):
            if line.strip() == '':
                continue

            m = uchk_re.fullmatch(line.strip())
            try:
                thp = m[UC_THP].split(':')
                if len(thp) != 3:
                    raise SyntaxError
            except TypeError:
                print(f"WARNING: command syntax error ({user_fp}:{fno})")
                continue
            except SyntaxError:
                print(f"WARNING: throughpoint syntax error ({user_fp}:{fno})")
                continue

            if thp[UCT_THP] != '':
                path_table = table.setdefault(thp[UCT_THP], {'all':{}})

                if thp[UCT_COR] != '':
                    corn_table = path_table.setdefault(thp[UCT_COR], {})
                else:
                    corn_table = path_table['all']

                if thp[UCT_PID] != '':
                    corn_table[int(thp[UCT_PID])] = (m[UC_STS], m[UC_COM])
                else:
                    corn_table['all'] = (m[UC_STS], m[UC_COM])

    return table
#}}}

def update_ucheck(wb: openpyxl.Workbook, uc_table: dict):
    """Update User Check to Report"""  #{{{
    for ws in wb.worksheets[1:]:
        for rid, thp in enumerate(ws.iter_rows(3, None, 7, 7, True), 3):
            if thp[0] in uc_table:
                path_table = uc_table[thp[0]]
                if ws.title in path_table:
                    corn_table = path_table[ws.title]
                else:
                    corn_table = path_table['all']

                pid = ws.cell(rid, 1).value
                if pid in corn_table:
                    ws.cell(rid, 6, corn_table[pid][UT_STS])
                    ws.cell(rid, 14, corn_table[pid][UT_COM])
                elif 'all' in corn_table:
                    ws.cell(rid, 6, corn_table['all'][UT_STS])
                    ws.cell(rid, 14, corn_table['all'][UT_COM])
#}}}

def summary_xls(wb: openpyxl.Workbook, fcnt_dict: dict, post_fp: str, pre_fp: str):
    """Summary"""  #{{{
    fcnt_dict = dict(sorted(fcnt_dict.items(), key=lambda x:x[1], reverse=True))
    wb._sheets = [wb._sheets[0]] + [wb._sheets[i] for i in fcnt_dict.keys()]

    status = ('Corner', 'Total\nNum', 'Total\nCheck', 'Fatal\nCheck', 
              'High\nWarn', 'Low\nWarn', 'User\nCheck')

    sum_ws = wb.worksheets[0]
    sum_ws.row_dimensions[1].height = 24
    sum_ws.column_dimensions['A'].width = 65

    rid = 1
    for cid, title in enumerate(status, 1):
        cell = sum_ws.cell(rid, cid, title)
        cell.fill = GREEN_FILL1
        cell.border = AR_BORDER1 
        cell.alignment = CC_ALIGN

    for sheet_name in wb.sheetnames[1:]:
        rid += 1
        sum_ws.cell(rid, 1, sheet_name)
        sum_ws.cell(rid, 2, f"='{sheet_name}'!A1")
        sum_ws.cell(rid, 3, f"='{sheet_name}'!B1")
        sum_ws.cell(rid, 4, f"='{sheet_name}'!C1")
        sum_ws.cell(rid, 5, f"='{sheet_name}'!D1")
        sum_ws.cell(rid, 6, f"='{sheet_name}'!E1")
        sum_ws.cell(rid, 7, f"='{sheet_name}'!F1")

        for cid in range(1, len(status)+1):
            cell = sum_ws.cell(rid, cid)
            cell.border = AR_BORDER2 
            cell.alignment = LC_ALIGN if cid == 1 else CC_ALIGN

    rid += 3
    cell = sum_ws.cell(rid, 1, f"postFlat report: {post_fp}")
    rid += 1
    cell = sum_ws.cell(rid, 1, f"preFlat report: {pre_fp}")

    for i, msg in enumerate(HELP_MSG, rid+2):
        sum_ws.cell(i, 1, msg)
#}}}

### Main ###

def create_argparse() -> argparse.ArgumentParser:
    """Create Argument Parser"""  #{{{
    parser = argparse.ArgumentParser(
                formatter_class=argparse.RawTextHelpFormatter,
                description="Compare table generator for IP flatten flow")

    parser.add_argument('-version', action='version', version=VERSION)
    parser.add_argument('-d', dest='is_dmsa', action='store_true', 
                                help="DMSA mode enable (default: single mode)")
    parser.add_argument('-c', dest='cfg_fn', metavar='<filename>', type=str, 
                                help=f"load custom module")
    parser.add_argument('-u', dest='user_fn', metavar='<filename>', type=str, 
                                help=f"user check setting")
    parser.add_argument('-r', dest='rpt_fn', metavar='<filename>', type=str, default=IP_FLAT_RPT_NAME, 
                                help=f"define report file name. (default: {IP_FLAT_RPT_NAME})")
    parser.add_argument('-f', dest='is_force', action='store_true', help="output force write")
    parser.add_argument('post_flat_dir', help="primetime report directory (post-flatten)") 
    parser.add_argument('pre_flat_dir', help="primetime report directory (pre-flatten)") 
    parser.add_argument('out_fn', help="output table name (format: xlsx)")

    return parser
#}}}

def main():
    """Main Function"""  #{{{
    parser = create_argparse()
    args = parser.parse_args()

    out_fn = args.out_fn if args.out_fn.endswith('.xlsx') else args.out_fn + '.xlsx'
    out_path = Path(out_fn)

    if out_path.exists() and not args.is_force:
        if input(f"{out_fn} existed, overwrite? (y/n) ").lower() != 'y':
            print('Terminal')
            return 

    if args.cfg_fn is not None:
        load_cfg(args.cfg_fn)

    post_dir = Path(args.post_flat_dir)
    pre_dir = Path(args.pre_flat_dir)
    corn_dict = {}

    if args.is_dmsa:
        for post_path in post_dir.glob(f'*/{args.rpt_fn}'):
            pre_path = pre_dir / post_path.parts[-2] / post_path.parts[-1]
            if not pre_path.is_file():
                print(f"WARNING: cannot find \'{pre_path}\', corner \'{post_path.parts[-2]}\' ignore")
            else:
                corn_dict[post_path.parts[-2]] = (post_path, pre_path)
    else:
        for post_path in [post_dir / args.rpt_fn]:
            pre_path = pre_dir / args.rpt_fn
            if not pre_path.is_file():
                print(f"WARNING: cannot find \'{pre_path}\'")
            else:
                corn_dict['single'] = (post_path, pre_path)

    wb = openpyxl.Workbook()
    wb.worksheets[0].title = 'summary'
    tid, fcnt_dict = 0, {}

    for corner, rpt_paths in corn_dict.items():
        try:
            tab = CORNER_MAP[corner]
        except KeyError:
            tab = corner

        col_sz = [0, 0]
        post_rpt = load_timing_report(rpt_paths[0], col_sz) 
        pre_rpt = load_timing_report(rpt_paths[1], col_sz)

        ws = wb.create_sheet(title=tab)
        tid += 1
        fcnt_dict[tid] = report_dump_xls(ws, post_rpt, pre_rpt, corner)

    if args.user_fn is not None:
        uc_table = load_ucheck_cfg(args.user_fn)
        update_ucheck(wb, uc_table)

    summary_xls(wb, fcnt_dict, post_dir.resolve(), pre_dir.resolve())
    wb.save(out_path)
    print(f"[INFO] finish (number of corner: {len(corn_dict)})")
#}}}

if __name__ == '__main__':
    main()

